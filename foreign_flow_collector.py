"""
==============================================================================
  THU THẬP DỮ LIỆU GIAO DỊCH - THỊ TRƯỜNG CHỨNG KHOÁN VIỆT NAM
  Nguon: CafeF | Moi ma 3 dong: Khoi ngoai / Tu doanh / Gia (%)
  Don vi: Nghin co phieu (KL rong)
==============================================================================
  Cai dat : pip install requests openpyxl pandas
  Chay    : python foreign_flow_collector.py
==============================================================================
"""

import requests
import pandas as pd
import numpy as np
import time
import warnings
from datetime import datetime, timedelta
warnings.filterwarnings("ignore")

# ─── CAU HINH ────────────────────────────────────────────────────────────────

# Danh sach VN30 hieu luc tu 02/02/2026 (BCM bi loai, VPL them moi)
WATCHLIST = [
    "ACB", "BID", "BVH", "CTG", "FPT",
    "GAS", "GVR", "HDB", "HPG", "MBB",
    "MSN", "MWG", "PLX", "POW", "SAB",
    "SSB", "SSI", "STB", "TCB", "TPB",
    "VCB", "VHM", "VIB", "VIC", "VJC",
    "VNM", "VPB", "VPL", "VRE", "KDH",
]

LOOKBACK_DAYS = 20
DELAY_SEC     = 1.5
TIMEOUT_SEC   = 30
MAX_RETRY     = 3
OUTPUT_FILE   = datetime.now().strftime("%Y%m%d_%H%M_foreign_flow.xlsx")

# Nguong phat hien KL bat thuong: > 300 trieu CP (sau khi chia 1000 = 300,000 nghin CP)
# Neu vuot nguong nay, goc du lieu co the la gia tri VND → fallback sang ThongKeDL
KL_ABNORMAL_THRESHOLD = 150_000  # don vi: nghin co phieu

URL_FOREIGN     = "https://cafef.vn/du-lieu/Ajax/PageNew/DataHistory/GDKhoiNgoai.ashx"
URL_PROPRIETARY = "https://cafef.vn/du-lieu/Ajax/PageNew/DataHistory/GDTuDoanh.ashx"
URL_THONGKE     = "https://cafef.vn/du-lieu/Ajax/PageNew/DataHistory/ThongKeDL.ashx"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36 Edg/146.0.0.0",
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "vi,en;q=0.9",
    "Referer": "https://cafef.vn/",
}

# ─── HAM GET VOI RETRY ────────────────────────────────────────────────────────

def get_json(url, params, label=""):
    for attempt in range(1, MAX_RETRY + 1):
        try:
            r = requests.get(url, headers=HEADERS, params=params, timeout=TIMEOUT_SEC)
            r.raise_for_status()
            return r.json()
        except Exception as e:
            if attempt < MAX_RETRY:
                print("\n    [" + label + " retry " + str(attempt) + "/" + str(MAX_RETRY) + "]", end=" ", flush=True)
                time.sleep(5)
            else:
                print("\n    [" + label + " loi] " + str(e))
                return None

# ─── HAM LAY THONGKE (fallback KL) ──────────────────────────────────────────

def fetch_thongke(symbol, page_size=35):
    """
    Lay ChenhLechKL tu ThongKeDL lam fallback khi KL chinh bi sai (tra ve gia tri VND).
    ChenhLechKL = KLDatMua - KLDatBan, don vi CP (co the co dau phan cach '.').
    Tra ve dict: {datetime: float (nghin CP)}
    """
    params = {"Symbol": symbol, "StartDate": "", "EndDate": "",
              "PageIndex": 1, "PageSize": page_size}
    data = get_json(URL_THONGKE, params, "TK")
    if not data or not data.get("Success"):
        return {}

    items = data.get("Data", {}).get("Data", [])
    result = {}
    for item in items:
        try:
            date = datetime.strptime(item.get("Date", ""), "%d/%m/%Y")
        except:
            continue
        try:
            # ChenhLechKL co dang "18.305.018" hoac "-13.162.288" → bo dau phan cach
            raw = item.get("ChenhLechKL", "0") or "0"
            kl = float(raw.replace(".", "").replace(",", ".").strip())
            result[date] = kl / 1000  # quy ve nghin co phieu
        except:
            result[date] = np.nan
    return result


def _check_and_fix(net_dict, symbol, label):
    """
    Kiem tra tung ngay trong net_dict.
    Neu bat ky gia tri nao vuot KL_ABNORMAL_THRESHOLD → fetch ThongKeDL
    va thay the toan bo cac gia tri bat thuong bang ChenhLechKL.
    In canh bao [FIX] neu co sua.
    """
    abnormal_dates = [d for d, v in net_dict.items()
                      if not np.isnan(v) and abs(v) > KL_ABNORMAL_THRESHOLD]
    if not abnormal_dates:
        return net_dict

    print("\n    [FIX " + label + "] " + symbol + ": " + str(len(abnormal_dates)) +
          " ngay bat thuong (|KL|>" + str(KL_ABNORMAL_THRESHOLD//1000) + "trieu) → fallback ThongKeDL",
          end=" ", flush=True)

    fallback = fetch_thongke(symbol, page_size=len(net_dict) + 5)
    time.sleep(DELAY_SEC)

    fixed = 0
    for d in abnormal_dates:
        if d in fallback and not np.isnan(fallback[d]):
            net_dict[d] = fallback[d]
            fixed += 1
        else:
            net_dict[d] = np.nan  # khong co fallback → de NaN

    print("(da sua " + str(fixed) + "/" + str(len(abnormal_dates)) + ")", end=" ", flush=True)
    return net_dict

# ─── HAM LAY DU LIEU ──────────────────────────────────────────────────────────

def fetch_foreign(symbol, page_size=35):
    params = {"Symbol": symbol, "StartDate": "", "EndDate": "",
              "PageIndex": 1, "PageSize": page_size}
    data = get_json(URL_FOREIGN, params, "NN")
    if not data or not data.get("Success"):
        return None, None

    items = data.get("Data", {}).get("Data", [])
    if not items:
        return None, None

    net_dict, price_dict = {}, {}
    for item in items:
        try:
            date = datetime.strptime(item.get("Ngay", ""), "%d/%m/%Y")
        except:
            continue
        # Khoi luong rong, don vi nghin co phieu
        net_dict[date] = float(item.get("KLGDRong", 0)) / 1000
        # % thay doi gia
        thay_doi = item.get("ThayDoi", "")
        try:
            pct = float(thay_doi.split("(")[1].replace("%)", "").replace(",", ".").strip())
        except:
            pct = np.nan
        price_dict[date] = pct

    net_dict = _check_and_fix(net_dict, symbol, "NN")
    return net_dict, price_dict


def fetch_proprietary(symbol, page_size=35):
    params = {"Symbol": symbol, "StartDate": "", "EndDate": "",
              "PageIndex": 1, "PageSize": page_size}
    data = get_json(URL_PROPRIETARY, params, "TD")
    if not data or not data.get("Success"):
        return None

    inner = data.get("Data", {}).get("Data", {})
    items = inner.get("ListDataTudoanh", []) if isinstance(inner, dict) else inner
    if not items:
        return None

    net_dict = {}
    for item in items:
        try:
            date = datetime.strptime(item.get("Date", ""), "%d/%m/%Y")
        except:
            continue
        # KL rong tu doanh, don vi nghin co phieu
        net_dict[date] = (float(item.get("KLcpMua", 0)) - float(item.get("KlcpBan", 0))) / 1000

    net_dict = _check_and_fix(net_dict, symbol, "TD")
    return net_dict

# ─── HAM HO TRO ───────────────────────────────────────────────────────────────

def get_trading_dates(n_days):
    dates, d = [], datetime.today()
    while len(dates) < n_days:
        if d.weekday() < 5:
            dates.append(d.strftime("%Y-%m-%d"))
        d -= timedelta(days=1)
    return list(reversed(dates))


def collect_all(symbols, dates):
    date_idx  = pd.to_datetime(dates)
    page_size = len(dates) + 5
    print("\nThu thap | " + dates[0] + " -> " + dates[-1] + " | " + str(len(symbols)) + " ma x 3 dong\n")

    all_rows = {}
    for i, sym in enumerate(symbols, 1):
        print("  [" + str(i).zfill(2) + "/" + str(len(symbols)) + "] " + sym + " ...", end=" ", flush=True)

        foreign, price = fetch_foreign(sym, page_size)
        s = pd.Series(foreign).reindex(date_idx) if foreign else pd.Series([np.nan]*len(dates), index=date_idx)
        all_rows[(sym, "Khoi ngoai")] = s.values
        print("NN:" + str(s.notna().sum()), end=" ")
        time.sleep(DELAY_SEC)

        prop = fetch_proprietary(sym, page_size)
        s = pd.Series(prop).reindex(date_idx) if prop else pd.Series([np.nan]*len(dates), index=date_idx)
        all_rows[(sym, "Tu doanh")] = s.values
        print("TD:" + str(s.notna().sum()), end=" ")
        time.sleep(DELAY_SEC)

        s = pd.Series(price).reindex(date_idx) if price else pd.Series([np.nan]*len(dates), index=date_idx)
        all_rows[(sym, "Gia (%)")] = s.values
        print("Gia:" + str(s.notna().sum()))

    col_labels = [d[5:].replace("-", "/") for d in dates]
    index = pd.MultiIndex.from_tuples(all_rows.keys(), names=["Ma CK", "Loai"])
    return pd.DataFrame(list(all_rows.values()), index=index, columns=col_labels)

# ─── XUAT EXCEL ───────────────────────────────────────────────────────────────

def build_excel(df, dates, symbols, output):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "GD Khoi Ngoai - Tu Doanh"

    H_BG="1F4E79"; H_FG="FFFFFF"
    BUY_BG="C6EFCE"; BUY_FG="006100"
    SEL_BG="FFC7CE"; SEL_FG="9C0006"
    PRC_BG="EBF3FB"; SYM_BG="2E4057"
    TD_BG="F2F2F2"
    TOT_BG="FFF2CC"; TOT_FG="7F6000"

    thin   = Side(style="thin",   color="BFBFBF")
    medium = Side(style="medium", color="1F4E79")
    n_days = len(dates)

    # Tieu de
    last_col = get_column_letter(n_days + 3)
    ws.merge_cells("A1:" + last_col + "1")
    t = ws["A1"]
    t.value = ("GIAO DICH KHOI NGOAI & TU DOANH (nghin CP)  |  "
               "Nguon: CafeF  |  Cap nhat: " + datetime.today().strftime("%d/%m/%Y"))
    t.font = Font(name="Arial", bold=True, size=12, color=H_FG)
    t.fill = PatternFill("solid", fgColor=H_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Headers
    headers = ["Ma CK", "Loai"] + [d[5:].replace("-", "/") for d in dates] + ["TONG KL"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = Font(name="Arial", bold=True, size=9, color=H_FG)
        c.fill = PatternFill("solid", fgColor=H_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[2].height = 32

    num_fmt = '#,##0.00;[Red]-#,##0.00;"-"'
    pct_fmt = '+0.00%;[Red]-0.00%;"-"'

    ri = 3
    for sym in symbols:
        row_start = ri
        for row_type in ["Khoi ngoai", "Tu doanh", "Gia (%)"]:
            is_price = row_type == "Gia (%)"
            is_td    = row_type == "Tu doanh"
            is_last  = row_type == "Gia (%)"
            bot      = medium if is_last else thin

            sc = ws.cell(row=ri, column=1, value=sym if row_type == "Khoi ngoai" else "")
            sc.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
            sc.fill = PatternFill("solid", fgColor=SYM_BG)
            sc.alignment = Alignment(horizontal="center", vertical="center")
            sc.border = Border(left=thin, right=thin, top=thin, bottom=bot)

            label_map = {"Khoi ngoai": "Khoi ngoai", "Tu doanh": "Tu doanh", "Gia (%)": "Gia (%)"}
            lc = ws.cell(row=ri, column=2, value=label_map[row_type])
            if is_price:
                lc.fill = PatternFill("solid", fgColor=PRC_BG)
                lc.font = Font(name="Arial", size=9, italic=True, color="2E4057")
            elif is_td:
                lc.fill = PatternFill("solid", fgColor=TD_BG)
                lc.font = Font(name="Arial", size=9, color="595959")
            else:
                lc.fill = PatternFill("solid", fgColor="DEEAF1")
                lc.font = Font(name="Arial", bold=True, size=9, color="1F4E79")
            lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            lc.border = Border(left=thin, right=thin, top=thin, bottom=bot)

            try:
                row_data = df.loc[(sym, row_type)].values
            except:
                row_data = [np.nan] * n_days

            for ci, val in enumerate(row_data, start=3):
                cell = ws.cell(row=ri, column=ci)
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.border = Border(left=thin, right=thin, top=thin, bottom=bot)
                if is_price:
                    cell.number_format = pct_fmt
                    cell.fill = PatternFill("solid", fgColor=PRC_BG)
                    if not pd.isna(val):
                        cell.value = val / 100
                        cell.font = Font(name="Arial", size=9,
                                        color=BUY_FG if val > 0 else (SEL_FG if val < 0 else "000000"))
                    else:
                        cell.font = Font(name="Arial", size=9)
                else:
                    cell.number_format = num_fmt
                    bg = TD_BG if is_td else "FFFFFF"
                    if pd.isna(val):
                        cell.fill = PatternFill("solid", fgColor=bg)
                        cell.font = Font(name="Arial", size=9)
                    elif val > 0:
                        cell.value = round(val, 2)
                        cell.fill = PatternFill("solid", fgColor=BUY_BG)
                        cell.font = Font(name="Arial", size=9, color=BUY_FG)
                    elif val < 0:
                        cell.value = round(val, 2)
                        cell.fill = PatternFill("solid", fgColor=SEL_BG)
                        cell.font = Font(name="Arial", size=9, color=SEL_FG)
                    else:
                        cell.value = 0.0
                        cell.fill = PatternFill("solid", fgColor=bg)
                        cell.font = Font(name="Arial", size=9)

            bc = get_column_letter(3); ec = get_column_letter(n_days + 2)
            tc = ws.cell(row=ri, column=n_days + 3)
            tc.alignment = Alignment(horizontal="right", vertical="center")
            tc.border = Border(left=thin, right=thin, top=thin, bottom=bot)
            if is_price:
                tc.value = '=IFERROR(AVERAGE(' + bc + str(ri) + ':' + ec + str(ri) + '),' + '"")'
                tc.number_format = pct_fmt
                tc.fill = PatternFill("solid", fgColor=PRC_BG)
                tc.font = Font(name="Arial", bold=True, size=9, italic=True)
            else:
                tc.value = '=IFERROR(SUM(' + bc + str(ri) + ':' + ec + str(ri) + '),' + '"")'
                tc.number_format = num_fmt
                tc.fill = PatternFill("solid", fgColor=TOT_BG)
                tc.font = Font(name="Arial", bold=True, size=9, color=TOT_FG)

            ws.row_dimensions[ri].height = 17
            ri += 1

        ws.merge_cells("A" + str(row_start) + ":A" + str(ri-1))

    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 12
    for ci in range(3, n_days + 3):
        ws.column_dimensions[get_column_letter(ci)].width = 8
    ws.column_dimensions[get_column_letter(n_days + 3)].width = 10
    ws.freeze_panes = "C3"

    wb.save(output)
    print("\nDa luu: " + output)

# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    dates = get_trading_dates(LOOKBACK_DAYS)
    print("Ngay giao dich: " + dates[0] + "  ->  " + dates[-1] + " (" + str(len(dates)) + " ngay)")
    df = collect_all(WATCHLIST, dates)
    build_excel(df, dates, WATCHLIST, OUTPUT_FILE)
    print("\nHoan thanh! File: " + OUTPUT_FILE)

if __name__ == "__main__":
    main()
